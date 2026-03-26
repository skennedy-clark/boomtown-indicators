"""
fetchers/fetch_income.py
------------------------
Fetches ATO individual income data (Table 8) from data.gov.au.

Strategy:
  1. Call data.gov.au CKAN API to list all resources in the dataset
  2. Find Table 8 by matching resource name
  3. Download the xlsx — one national file serves all towns
  4. Extract values for each town by postcode

This means when a new year is published (e.g. 2023-24), you only need to
update LATEST_YEAR and DATASET_SLUG — the URL resolves automatically.

Table 8 content:
  Median and average taxable income by state/territory and postcode,
  2003-04 and 2013-14 to 2022-23 income years.
  This is the multi-year historical series the researchers use.

Confirmed working:
  Dataset : taxation-statistics-2022-23
  Resource: e4b8c6b4-1185-4be6-9bd7-cf4385934d00
  File    : ts23individual08medianaveragetaxableincomestatepostcode.xlsx
  Size    : ~2MB
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from fetchers.base import BaseFetcher

try:
    import pandas as pd
    import requests
except ImportError:
    raise ImportError("pip install pandas openpyxl requests")


# ── Configuration ──────────────────────────────────────────────────────────────

LATEST_YEAR  = "2022-23"

# data.gov.au dataset slugs by year — update when ATO publishes new year
DATASET_SLUGS = {
    "2022-23": "taxation-statistics-2022-23",
    "2021-22": "taxation-statistics-2021-22",
    "2020-21": "taxation-statistics-2020-21",
}

# CKAN API endpoint — note: data.gov.au uses /data/api/ not /api/
CKAN_API = "https://data.gov.au/data/api/3/action/package_show?id={slug}"

# How to identify Table 8 among the dataset's resources
# We match on resource name (case-insensitive)
TABLE8_NAME_FRAGMENTS = ["table 8", "individual", "postcode", "median"]


class ATOIncomeFetcher(BaseFetcher):

    SOURCE_NAME      = "ato_income"
    SUPPORTED_STATES = []   # ATO covers all states

    def fetch_all(self):
        year = LATEST_YEAR
        slug = DATASET_SLUGS.get(year)
        if not slug:
            self.result.add_error("ALL", f"No dataset slug configured for year {year}")
            return

        # ── Step 1: Resolve Table 8 URL via CKAN API ──────────────────────────
        t8_url = self._find_table8_url(slug, year)
        if not t8_url:
            return  # errors already logged

        # ── Step 2: Download (cached after first run) ─────────────────────────
        cache_key = f"ato_table8_{year}"
        t8_path   = self.download(t8_url, cache_key, suffix=".xlsx")
        if t8_path is None:
            self.result.add_error("ALL", f"Download failed for Table 8 ({year})")
            return

        # ── Step 3: Verify file is a real xlsx ────────────────────────────────
        if not self._verify_xlsx(t8_path):
            self.log.error(
                f"Downloaded file is not a valid xlsx "
                f"({t8_path.stat().st_size} bytes) — invalidating cache"
            )
            self.cache.invalidate(cache_key)
            t8_path.unlink(missing_ok=True)
            self.result.add_error("ALL", "Table 8 download invalid — will retry next run")
            return

        # ── Step 4: Parse into postcode lookup ────────────────────────────────
        self.log.info(f"Parsing {t8_path.name} ({t8_path.stat().st_size / 1024:.0f} KB)")
        postcode_data = self._parse_table8(t8_path)
        if not postcode_data:
            self.result.add_error("ALL", "Table 8 parsed empty — check file structure")
            return

        # ── Step 5: Extract per-town ──────────────────────────────────────────
        for town in self.applicable_towns():
            self._extract_town(town, postcode_data, year)

    # ── CKAN URL resolution ────────────────────────────────────────────────────

    def _find_table8_url(self, slug: str, year: str) -> str | None:
        """
        Query the CKAN API for the dataset, then find the Table 8 resource
        by matching name fragments. Returns the download URL or None.
        """
        api_url = CKAN_API.format(slug=slug)
        self.log.info(f"Querying CKAN API: {api_url}")

        try:
            resp = requests.get(api_url, timeout=15)
            resp.raise_for_status()
            data = resp.json()

            if not data.get("success"):
                self.log.error(f"CKAN API returned success=false: {data.get('error')}")
                self.result.add_error("ALL", f"CKAN API error for {slug}")
                return None

            resources = data["result"]["resources"]
            self.log.info(f"  Dataset has {len(resources)} resources")

            # Find Table 8 — score each resource by how many fragments match
            best_url   = None
            best_score = 0

            for r in resources:
                name = r.get("name", "").lower()
                url  = r.get("url", "")
                score = sum(1 for f in TABLE8_NAME_FRAGMENTS if f in name)
                if score > best_score:
                    best_score = score
                    best_url   = url
                    self.log.debug(f"  Candidate (score {score}): {name} -> {url}")

            if best_url and best_score >= 2:
                self.log.info(f"  Table 8 URL (score {best_score}): {best_url}")
                return best_url

            # Fallback: search by filename fragment in URL
            for r in resources:
                url = r.get("url", "")
                if "individual08" in url.lower() and "postcode" in url.lower():
                    self.log.info(f"  Table 8 URL (filename match): {url}")
                    return url

            self.log.error(
                f"Could not find Table 8 among {len(resources)} resources. "
                f"First 5 resource names: "
                f"{[r.get('name','') for r in resources[:5]]}"
            )
            self.result.add_error("ALL", f"Table 8 not found in dataset {slug}")
            return None

        except requests.RequestException as exc:
            self.log.error(f"CKAN API request failed: {exc}")
            self.result.add_error("ALL", f"CKAN API unreachable: {exc}")
            return None

    # ── Validation ─────────────────────────────────────────────────────────────

    def _verify_xlsx(self, path: Path) -> bool:
        """Check file is a real xlsx — must be >100KB and start with PK magic bytes."""
        if path.stat().st_size < 100_000:
            self.log.warning(f"File too small: {path.stat().st_size} bytes (expected >100KB)")
            return False
        # xlsx files are zip archives — start with PK\x03\x04
        magic = path.read_bytes()[:4]
        if magic != b"PK\x03\x04":
            self.log.warning(f"File does not have xlsx/zip magic bytes: {magic!r}")
            return False
        return True

    # ── Parser ─────────────────────────────────────────────────────────────────

    def _parse_table8(self, path: Path) -> dict:
        """
        Table 8 actual structure (confirmed from file inspection):
          Row 1 : Title text (skip)
          Row 2 : Column headers — embed year in name, e.g.:
                  "Average3 taxable income 2022-23 \n$"
          Row 3+: Data — State, Postcode (int), then numeric values

        We extract only the "Average taxable income" columns since that
        is what the booklets report. Keys are normalised to "YYYY-YY".

        Returns:
          { '4455': { '2003-04': 38500.0, '2022-23': 68400.0, ... } }
        """
        try:
            import openpyxl as _openpyxl
            wb = _openpyxl.load_workbook(path, read_only=True)
            self.log.info(f"  Sheets: {wb.sheetnames}")

            # Find the right sheet
            ws = None
            for name in wb.sheetnames:
                if "table 8" in name.lower():
                    ws = wb[name]
                    self.log.info(f"  Using sheet: '{name}'")
                    break
            if ws is None:
                skip = {"notes", "contents", "readme", "index"}
                for name in wb.sheetnames:
                    if name.lower() not in skip:
                        ws = wb[name]
                        self.log.info(f"  Fallback sheet: '{name}'")
                        break

            if ws is None:
                self.log.error("No usable sheet found")
                return {}

            rows = list(ws.iter_rows(values_only=True))
            # Row 0 = title, Row 1 = headers, Row 2+ = data
            header_row = rows[1]

            # Find postcode column index
            pc_idx = next(
                (i for i, h in enumerate(header_row)
                 if h and "postcode" in str(h).lower()),
                None
            )
            if pc_idx is None:
                self.log.error(f"No postcode column. Headers: {header_row[:6]}")
                return {}

            # Find "Average taxable income" columns; extract year from header text
            # e.g. "Average3 taxable income 2022-23 \n$"
            year_pat = re.compile(r"(\d{4}[\u2013\u002d]\d{2})")
            avg_cols = {}  # { col_index: "2022-23" }
            for i, h in enumerate(header_row):
                if h and "average" in str(h).lower() and "taxable" in str(h).lower():
                    m = year_pat.search(str(h))
                    if m:
                        yr = m.group(1).replace("\u2013", "-")
                        avg_cols[i] = yr

            if not avg_cols:
                self.log.error(
                    f"No average taxable income columns found. "
                    f"Sample headers: {[h for h in header_row[:8] if h]}"
                )
                return {}

            years = sorted(set(avg_cols.values()))
            self.log.info(
                f"  Year range: {years[0]} to {years[-1]} "
                f"({len(years)} years)"
            )

            result  = {}
            skipped = 0
            for row in rows[2:]:
                pc_raw = row[pc_idx]
                if pc_raw is None:
                    skipped += 1
                    continue
                pc = str(int(pc_raw)) if isinstance(pc_raw, float) else str(pc_raw).strip()
                pc = pc.replace(".0", "")
                if not pc.isdigit() or not (3 <= len(pc) <= 4):
                    skipped += 1
                    continue
                pc = pc.zfill(4)
                year_values = {}
                for col_i, yr in avg_cols.items():
                    val = row[col_i]
                    if val is not None:
                        try:
                            year_values[yr] = float(val)
                        except (ValueError, TypeError):
                            pass
                if year_values:
                    result[pc] = year_values

            self.log.info(
                f"  Parsed {len(result)} postcodes "
                f"(skipped {skipped} non-postcode rows)"
            )
            return result

        except Exception as exc:
            self.log.error(f"Table 8 parse error: {exc}", exc_info=True)
            return {}

    # ── Per-town extraction ────────────────────────────────────────────────────

    def _extract_town(self, town, postcode_data: dict, year: str):
        if not town.postcodes:
            self.result.add_warning(town.name, "No postcodes configured — skipping")
            self.result.towns_skipped.append(town.name)
            return

        combined_years: dict[str, list[float]] = {}
        found_any = False

        for pc in town.postcodes:
            pc_str = str(pc).zfill(4)
            if pc_str in postcode_data:
                found_any = True
                for yr, val in postcode_data[pc_str].items():
                    combined_years.setdefault(yr, []).append(val)
            else:
                self.result.add_warning(
                    town.name,
                    f"Postcode {pc_str} not in Table 8 "
                    f"(suppressed by ATO or not in dataset)"
                )

        if not found_any:
            self.result.add_error(
                town.name, "No income data found for any postcode"
            )
            self.result.towns_failed.append(town.name)
            return

        # Average across postcodes where a town spans multiple
        avg_by_year = {
            yr: sum(vals) / len(vals)
            for yr, vals in sorted(combined_years.items())
        }

        n_years = len(avg_by_year)
        latest_val = avg_by_year.get(year, "N/A")

        out = {
            "town":     town.name,
            "state":    town.state,
            "postcodes": town.postcodes,
            "source":   "ATO Taxation Statistics Table 8",
            "latest_year": year,
            "avg_taxable_income_by_year": avg_by_year,
        }

        out_dir  = Path(__file__).parent.parent / "cache" / "ato"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{town.slug}_income.json"

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, default=str)

        self.log.info(
            f"  {town.name}: {n_years} years, "
            f"latest ({year}) = ${latest_val:,.0f} "
            f"-> {out_path.name}"
            if isinstance(latest_val, float)
            else f"  {town.name}: {n_years} years -> {out_path.name}"
        )
        self.result.towns_ok.append(town.name)
        self.result.cached_files.append(out_path)


if __name__ == "__main__":
    from logger import get_logger
    get_logger()
    result = ATOIncomeFetcher().run()
    sys.exit(0 if result.success else 1)