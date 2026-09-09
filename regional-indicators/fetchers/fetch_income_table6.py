"""
fetchers/fetch_income_table6.py
--------------------------------
Fetches ATO individual income data (Table 6) from data.gov.au.

Table 6 is published annually with two sheets:
  Table 6A — rows split by taxable status (Non Taxable / Taxable)
  Table 6B — combined totals (both statuses merged, no status column)

We use:
  Table 6A (Taxable rows only) → avg_income_taxable
    = taxable_income_$ / taxable_income_no.
  Table 6B (combined)          → earners_no, wages_total

Confirmed column indices (2022-23 file):
  6A: 0=status  3=postcode  5=taxable_income_no  6=taxable_income_$
      19=wages_no  20=wages_$
  6B: 2=postcode  18=wages_no  19=wages_$

Confirmed values Roma (4455) 2022-23:
  avg_income_taxable = $82,899    ✓ ref: Income - for taxable individuals.csv
  earners_no         = 4,515      ✓ ref: Number of earners.csv
  wages_total        = $312,571,660  ✓ ref: Wage & salary earnings (town total).csv

Website CSVs produced:
  Income - for taxable individuals.csv
  Number of earners.csv
  Wage & salary earnings (town total).csv
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from fetchers.base import BaseFetcher
from fetchers.ato_release import discover_latest_ato_release

try:
    import openpyxl
    import requests
except ImportError:
    raise ImportError("pip install openpyxl requests")


CKAN_API = "https://data.gov.au/data/api/3/action/package_show"


def _score_resource(name: str) -> int:
    n = name.lower()
    score = 0
    if "table" in n and ("6" in n or "06" in n): score += 3
    if "postcode" in n:                           score += 2
    if "individual" in n:                         score += 1
    if "selected" in n and "item" in n:           score += 1
    if "taxable" in n and "status" in n:          score += 1
    return score


class ATOTable6Fetcher(BaseFetcher):

    SOURCE_NAME      = "ato_income_table6"
    SUPPORTED_STATES = []

    def fetch_all(self):
        try:
            release = discover_latest_ato_release()
        except RuntimeError as exc:
            self.result.add_error("ALL", str(exc))
            return

        year = release.financial_year
        slug = release.slug

        self.log.info(
            f"Using ATO Taxation Statistics {year} "
            f"(package modified {release.modified or 'unknown'})"
        )

        try:
            resp = requests.get(CKAN_API, params={"id": slug}, timeout=30)
            resp.raise_for_status()
            resources = resp.json().get("result", {}).get("resources", [])
        except Exception as exc:
            self.log.error(f"CKAN API unreachable: {exc}")
            self.result.add_error("ALL", f"CKAN API unreachable: {exc}")
            return

        self.log.info(f"  Dataset has {len(resources)} resources")
        best = max(resources, key=lambda r: _score_resource(r.get("name", "")), default=None)
        if not best or _score_resource(best.get("name", "")) == 0:
            self.log.error("Could not identify Table 6 resource")
            return

        t6_url = best["url"]
        self.log.info(f"  Table 6 URL: {t6_url}")

        cache_key = f"ato_table6_{year}"
        t6_path   = self.download(t6_url, cache_key, suffix=".xlsx")
        if not t6_path:
            self.result.add_error("ALL", "Table 6 download failed")
            return

        self.log.info(f"  Parsing {t6_path.name} ({t6_path.stat().st_size // 1024} KB)")

        taxable_data  = self._parse_6a_taxable(t6_path)
        combined_data = self._parse_6b_combined(t6_path)

        if not taxable_data or not combined_data:
            self.result.add_error("ALL", "Table 6 parse returned no data")
            return

        for town in self.applicable_towns():
            self._extract_town(town, taxable_data, combined_data, year)

    def _parse_6a_taxable(self, path: Path) -> dict:
        """Table 6A — Taxable rows only. Returns {pc: {taxable_no, taxable_income}}"""
        try:
            ws   = openpyxl.load_workbook(path, read_only=True)["Table 6A"]
            rows = list(ws.iter_rows(values_only=True))
            result, skipped = {}, 0
            for row in rows[2:]:
                status = str(row[0] or "").strip().lower()
                if "taxable" not in status or "non" in status:
                    continue
                pc = self._clean_pc(row[3])
                if not pc:
                    skipped += 1
                    continue
                try:
                    result[pc] = {
                        "taxable_no":     float(row[5] or 0),
                        "taxable_income": float(row[6] or 0),
                    }
                except (TypeError, ValueError):
                    skipped += 1
            self.log.info(f"    6A: {len(result)} taxable postcodes (skipped {skipped})")
            return result
        except Exception as exc:
            self.log.error(f"Table 6A parse error: {exc}", exc_info=True)
            return {}

    def _parse_6b_combined(self, path: Path) -> dict:
        """Table 6B — combined totals. Returns {pc: {wages_no, wages_total}}"""
        try:
            ws   = openpyxl.load_workbook(path, read_only=True)["Table 6B"]
            rows = list(ws.iter_rows(values_only=True))
            result, skipped = {}, 0
            for row in rows[2:]:
                pc = self._clean_pc(row[2])
                if not pc:
                    skipped += 1
                    continue
                try:
                    result[pc] = {
                        "wages_no":    float(row[18] or 0),
                        "wages_total": float(row[19] or 0),
                    }
                except (TypeError, ValueError):
                    skipped += 1
            self.log.info(f"    6B: {len(result)} postcodes (skipped {skipped})")
            return result
        except Exception as exc:
            self.log.error(f"Table 6B parse error: {exc}", exc_info=True)
            return {}

    def _clean_pc(self, raw) -> str | None:
        if raw is None:
            return None
        pc = str(int(raw)) if isinstance(raw, float) else str(raw).strip()
        pc = pc.replace(".0", "").zfill(4)
        return pc if (pc.isdigit() and len(pc) == 4) else None

    def _extract_town(self, town, taxable_data: dict, combined_data: dict, year: str):
        agg_t = {"taxable_no": 0.0, "taxable_income": 0.0}
        agg_c = {"wages_no": 0.0, "wages_total": 0.0}
        found = False

        for pc in town.postcodes:
            key = pc.zfill(4)
            if key in taxable_data:
                found = True
                for k in agg_t:
                    agg_t[k] += taxable_data[key][k]
            if key in combined_data:
                for k in agg_c:
                    agg_c[k] += combined_data[key][k]

        if not found:
            self.log.warning(f"  [{town.name}] no Table 6 data for {town.postcodes}")
            self.result.towns_failed.append(town.name)
            return

        avg_taxable = (
            round(agg_t["taxable_income"] / agg_t["taxable_no"])
            if agg_t["taxable_no"] > 0 else None
        )
        earners_no  = int(agg_c["wages_no"])
        wages_total = int(agg_c["wages_total"])

        fy_parts = year.replace("\u2013", "-").split("-")
        cal_year = str(int(fy_parts[0]) + 1) if len(fy_parts) == 2 else year

        out = {
            "town": town.name, "state": town.state, "postcodes": town.postcodes,
            "source": "ATO Taxation Statistics Table 6",
            "latest_year": year, "cal_year": cal_year,
            "indicators": {
                "avg_income_taxable": {cal_year: avg_taxable},
                "earners_no":         {cal_year: earners_no},
                "wages_total":        {cal_year: wages_total},
            }
        }

        out_dir  = Path(__file__).parent.parent / "cache" / "ato"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{town.slug}_income_t6.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, default=str)

        self.log.info(
            f"  {town.name}: avg_taxable=${avg_taxable:,}  "
            f"earners={earners_no:,}  wages=${wages_total:,}"
            if avg_taxable else f"  {town.name}: no data"
        )
        self.result.towns_ok.append(town.name)


if __name__ == "__main__":
    from logger import get_logger
    get_logger()
    result = ATOTable6Fetcher().run()
    sys.exit(0 if result.success else 1)
