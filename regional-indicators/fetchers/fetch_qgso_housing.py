"""
fetchers/fetch_qgso_housing.py
-------------------------------
Parses the QGSO + BOM annual data file for housing and NRW indicators.

*** MANUAL DOWNLOAD REQUIRED ***
This file is assembled by the researcher from QGSO QRIS each year.
It cannot be automated because QGSO QRIS requires a web form login.

Setup each year:
  1. Log into QGSO QRIS and download the housing / NRW data
  2. Assemble into the standard xlsx format (see existing file as template)
  3. Save as: cache/qgso_and_bom_{year}.xlsx
  4. Run: python run_update.py --only qgso_housing

What this fetcher extracts:
  - Housing: house sale price (median), house sales (number)
  - Housing: rent (3-bedroom median), residential building approvals
  - NRW: non-resident workers on-shift (UCL level and LGA level)

What is NOT in this fetcher (use separate automated fetchers instead):
  - Unemployment rate → fetch_salm_unemployment.py (DEWR SALM, fully automated)
  - Rainfall → fetch_bom_rainfall.py (BOM Climate Data Online, fully automated)

Source: QGSO QRIS — Residential Land and Dwelling Sales, Rental Bond data
        (QLD only; data is not publicly available for download without login)

SA2 mapping:
  The QGSO file uses ASGS Edition 2 SA2 codes, stored in towns.toml as
  qgso_sa2. These differ from the Edition 3 codes in sa2_code.

Aggregation methods (confirmed against reference CSVs):
  house_sales : sum of 4 rolling quarterly totals
                e.g. Roma 2024: 184 + 199 + 206 + 206 = 795
  house_price : mean of 4 quarterly medians
                e.g. Roma 2024: (325k+340k+345k+350k)/4 = $340,000
  rent        : mean of 4 quarterly medians
                e.g. Roma 2024: (360+360+373+380)/4 = $368.25
  approvals   : single annual value (year ended Dec)

Website CSVs produced:
  House sale price (median).csv
  House sales (number).csv
  Rent (3-bedroom house; median).csv
  Residential building approvals.csv
  Non-resident workers in town.csv
  Non-resident workers in local govt area.csv
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from fetchers.base import BaseFetcher
from config import CACHE_DIR

try:
    import openpyxl
except ImportError:
    raise ImportError("pip install openpyxl")


YEAR      = 2024
CACHE_KEY = f"qgso_and_bom_{YEAR}"
CACHE_FILE = CACHE_DIR / f"{CACHE_KEY}.xlsx"


class QGSOHousingFetcher(BaseFetcher):

    SOURCE_NAME      = "qgso_housing"
    SUPPORTED_STATES = ["QLD"]

    def fetch_all(self):
        if not CACHE_FILE.exists():
            self.log.error(
                f"File not found: {CACHE_FILE}\n"
                f"  Manual step required:\n"
                f"  1. Download housing/NRW data from QGSO QRIS\n"
                f"  2. Assemble into standard xlsx format\n"
                f"  3. Save as: {CACHE_FILE}"
            )
            self.result.add_error("ALL", f"Missing: {CACHE_FILE.name}")
            return

        self.log.info(f"  Reading {CACHE_FILE.name} ({CACHE_FILE.stat().st_size // 1024} KB)")

        wb = openpyxl.load_workbook(CACHE_FILE, read_only=True)

        house_price, house_sales = self._parse_sales(wb)
        rent                     = self._parse_rent(wb)
        approvals                = self._parse_approvals(wb)
        nrw_town, nrw_lga        = self._parse_nrw(wb)

        out_dir = CACHE_DIR / "housing"
        out_dir.mkdir(exist_ok=True)

        for town in self.applicable_towns():
            self._extract_town(
                town, house_price, house_sales, rent, approvals,
                nrw_town, nrw_lga, out_dir
            )

    def _parse_sales(self, wb) -> tuple[dict, dict]:
        """
        Returns (median_price, num_sales) each { sa2_key: value }
        house_sales : sum of 4 rolling quarterly totals
        house_price : mean of 4 quarterly medians
        """
        ws   = wb['Sales']
        rows = list(ws.iter_rows(values_only=True))

        series_blocks = []
        current_label = ""
        for i, row in enumerate(rows):
            if row[0] and 'SERIES' in str(row[0]):
                current_label = str(row[0])
            if row[0] == 'Region':
                series_blocks.append((i, current_label))

        price_data = {}
        sales_data = {}

        for header_idx, label in series_blocks:
            header   = rows[header_idx]
            is_price = 'median sale price' in label.lower() or 'median' in label.lower()
            data_cols = [i for i, v in enumerate(header) if v and i > 0]

            for row in rows[header_idx + 1:]:
                if not row[0]:
                    continue
                if row[0] == 'Region' or 'SERIES' in str(row[0]):
                    break
                key = str(row[0]).split(' - ')[0].strip()
                if 'SA2/' not in key and 'LGA/' not in key:
                    continue
                vals = []
                for col_i in data_cols:
                    v = row[col_i]
                    if v is not None:
                        try:
                            vals.append(float(v))
                        except (TypeError, ValueError):
                            pass
                if vals:
                    if is_price:
                        price_data[key] = round(sum(vals) / len(vals))
                    else:
                        sales_data[key] = int(sum(vals))

        self.log.info(f"  Sales: {len(price_data)} price, {len(sales_data)} count")
        return price_data, sales_data

    def _parse_rent(self, wb) -> dict:
        """Returns { sa2_key: median_rent } — mean of 4 quarterly medians."""
        ws   = wb['rent']
        rows = list(ws.iter_rows(values_only=True))
        header   = rows[0]
        last_col = max(i for i, v in enumerate(header) if v and i > 0)

        result = {}
        for row in rows[1:]:
            if not row[0]:
                continue
            key = str(row[0]).split(' - ')[0].strip()
            if 'SA2/' not in key and 'LGA/' not in key:
                continue
            vals = [float(row[i]) for i in range(1, last_col + 1) if row[i] is not None]
            if vals:
                result[key] = round(sum(vals) / len(vals), 2)

        self.log.info(f"  Rent: {len(result)} series parsed")
        return result

    def _parse_approvals(self, wb) -> dict:
        """Returns { sa2_key: residential_approvals } — single annual value."""
        ws   = wb['Building approvals']
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(i for i, r in enumerate(rows) if r[0] == 'Region')

        result = {}
        for row in rows[header_idx + 1:]:
            if not row[0]:
                continue
            key = str(row[0]).split(' - ')[0].strip()
            if 'SA2/' not in key and 'LGA/' not in key:
                continue
            val = row[1]
            if val is not None:
                try:
                    result[key] = int(float(val))
                except (TypeError, ValueError):
                    pass

        self.log.info(f"  Approvals: {len(result)} regions parsed")
        return result

    def _parse_nrw(self, wb) -> tuple[dict, dict]:
        """
        Returns (nrw_town, nrw_lga)
        nrw_town: { ucl_name: on_shift_count }
        nrw_lga:  { lga_name: on_shift_count }
        """
        ws   = wb['NRW']
        rows = list(ws.iter_rows(values_only=True))

        nrw_town = {}
        nrw_lga  = {}
        current_lga = None

        for row in rows[2:]:
            lga, location, ucl, erp, nrw, fte = (
                row[0], row[1], row[2], row[3], row[4], row[5]
            )
            if lga:
                current_lga = str(lga).strip()

            loc_str = str(location or '').lower()
            if ucl and nrw is not None and 'total' not in loc_str and 'other towns' not in str(ucl).lower():
                try:
                    nrw_town[str(ucl).strip()] = int(float(nrw))
                except (TypeError, ValueError):
                    pass

            if location and 'total' in str(location or '').lower() and current_lga and nrw:
                try:
                    nrw_lga[current_lga] = int(float(nrw))
                except (TypeError, ValueError):
                    pass

        self.log.info(f"  NRW: {len(nrw_town)} UCL, {len(nrw_lga)} LGA")
        return nrw_town, nrw_lga

    def _extract_town(self, town, house_price, house_sales, rent, approvals,
                      nrw_town, nrw_lga, out_dir):
        sa2_key = f"SA2/{town.qgso_sa2}"
        lga_key = town.qgso_lga

        def get(d, *keys):
            for k in keys:
                if k and k in d:
                    return d[k]
            return None

        yr = str(YEAR)
        indicators = {
            "house_price":  get(house_price, sa2_key),
            "house_sales":  get(house_sales, sa2_key),
            "rent":         get(rent, sa2_key),
            "approvals":    get(approvals, sa2_key, lga_key),
            "nrw_town":     nrw_town.get(town.name),
            "nrw_lga":      nrw_lga.get(town.lga),
        }

        found = {k: v for k, v in indicators.items() if v is not None}
        if not found:
            self.log.warning(f"  [{town.name}] no housing/NRW data found")
            self.result.towns_failed.append(town.name)
            return

        out = {
            "town":      town.name,
            "state":     town.state,
            "qgso_sa2":  town.qgso_sa2,
            "qgso_lga":  town.qgso_lga,
            "year":      YEAR,
            "source":    "QGSO QRIS (manual download — housing and NRW)",
            "indicators": {k: {yr: v} for k, v in found.items()},
        }

        out_path = out_dir / f"{town.slug}_qgso_{YEAR}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2)

        summary = ", ".join(f"{k}={v}" for k, v in found.items())
        self.log.info(f"  {town.name}: {summary}")
        self.result.towns_ok.append(town.name)


if __name__ == "__main__":
    from logger import get_logger
    get_logger()
    result = QGSOHousingFetcher().run()
    sys.exit(0 if result.success else 1)